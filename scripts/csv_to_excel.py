#!/usr/bin/env python3
"""Convert MUIS CSV exports to Excel format.

This script converts MUIS CSV files (89 columns, 3-row header) to Excel format
with proper formatting:
- Preserves 3-row header structure (metadata, column names, validation rules)
- Freezes top 3 rows for easier scrolling
- Auto-adjusts column widths for readability
- Bolds header rows
- Preserves UTF-8 encoding (Estonian characters: õ, ä, ö, ü)

Usage:
    python scripts/csv_to_excel.py input.csv output.xlsx
    python scripts/csv_to_excel.py input.csv  # Auto-generates output filename
"""

import argparse
import csv
import sys
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def csv_to_excel(
    csv_path: Path,
    excel_path: Optional[Path] = None,
    freeze_header_rows: int = 3,
    auto_width: bool = True,
) -> Path:
    """Convert MUIS CSV to Excel with formatting.

    Args:
        csv_path: Path to input CSV file
        excel_path: Path to output Excel file (optional, auto-generated if None)
        freeze_header_rows: Number of header rows to freeze (default: 3)
        auto_width: Whether to auto-adjust column widths (default: True)

    Returns:
        Path to created Excel file

    Raises:
        FileNotFoundError: If CSV file doesn't exist
        ValueError: If CSV is empty or malformed
    """
    csv_path = Path(csv_path)
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    # Auto-generate Excel path if not provided
    if excel_path is None:
        excel_path = csv_path.with_suffix(".xlsx")
    else:
        excel_path = Path(excel_path)

    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MUIS Export"

    # Read CSV and write to Excel
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows_written = 0

        for row_idx, row in enumerate(reader, start=1):
            if not row:
                continue

            # Write row to Excel
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format header rows (first 3 rows)
                if row_idx <= freeze_header_rows:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="top")

            rows_written += 1

    if rows_written == 0:
        raise ValueError(f"CSV file is empty: {csv_path}")

    # Freeze header rows
    if freeze_header_rows > 0:
        # Freeze at row after header rows
        freeze_cell = f"A{freeze_header_rows + 1}"
        ws.freeze_panes = freeze_cell

    # Auto-adjust column widths
    if auto_width:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    # Calculate length, add padding
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            # Set adjusted width (add padding, cap at reasonable max)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(excel_path)
    return excel_path


def main() -> int:
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Convert MUIS CSV exports to Excel format",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s input.csv output.xlsx
  %(prog)s vabamu_sample_100_v2.1.csv  # Auto-generates .xlsx filename
  %(prog)s --no-freeze input.csv output.xlsx
        """,
    )

    parser.add_argument(
        "csv_file",
        type=Path,
        help="Input CSV file path",
    )

    parser.add_argument(
        "excel_file",
        type=Path,
        nargs="?",
        default=None,
        help="Output Excel file path (optional, auto-generated if not provided)",
    )

    parser.add_argument(
        "--no-freeze",
        action="store_true",
        help="Don't freeze header rows",
    )

    parser.add_argument(
        "--no-auto-width",
        action="store_true",
        help="Don't auto-adjust column widths",
    )

    args = parser.parse_args()

    try:
        output_path = csv_to_excel(
            csv_path=args.csv_file,
            excel_path=args.excel_file,
            freeze_header_rows=0 if args.no_freeze else 3,
            auto_width=not args.no_auto_width,
        )

        print(f"✅ Converted: {args.csv_file}")
        print(f"📊 Excel output: {output_path}")
        return 0

    except FileNotFoundError as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        return 1

    except ValueError as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        return 1

    except Exception as e:
        print(f"❌ Unexpected error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
