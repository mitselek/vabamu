"""Tests for CSV to Excel converter."""

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scripts.csv_to_excel import csv_to_excel


class TestCsvToExcel:
    """Tests for csv_to_excel function."""

    def test_basic_conversion(self, tmp_path: Path) -> None:
        """Should convert CSV to Excel with basic structure."""
        # Create test CSV
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Header1", "Header2", "Header3"])
            writer.writerow(["Value1", "Value2", "Value3"])
            writer.writerow(["Value4", "Value5", "Value6"])

        # Convert to Excel
        excel_file = tmp_path / "test.xlsx"
        result_path = csv_to_excel(csv_file, excel_file)

        # Verify file created
        assert result_path.exists()
        assert result_path == excel_file

        # Verify structure
        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.max_row == 3
        assert ws.max_column == 3

        # Verify data
        assert ws.cell(1, 1).value == "Header1"
        assert ws.cell(2, 1).value == "Value1"
        assert ws.cell(3, 3).value == "Value6"

        wb.close()

    def test_auto_filename_generation(self, tmp_path: Path) -> None:
        """Should auto-generate Excel filename if not provided."""
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["A", "B"])

        # Convert without specifying Excel path
        result_path = csv_to_excel(csv_file)

        # Should create .xlsx with same basename
        expected_path = tmp_path / "test.xlsx"
        assert result_path == expected_path
        assert result_path.exists()

    def test_freeze_panes(self, tmp_path: Path) -> None:
        """Should freeze header rows by default."""
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for i in range(5):
                writer.writerow([f"Row{i}Col{j}" for j in range(3)])

        excel_file = csv_to_excel(csv_file, freeze_header_rows=3)

        # Check freeze panes
        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.freeze_panes == "A4"  # Frozen at row 4
        wb.close()

    def test_no_freeze_panes(self, tmp_path: Path) -> None:
        """Should not freeze panes if freeze_header_rows=0."""
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["A", "B"])

        excel_file = csv_to_excel(csv_file, freeze_header_rows=0)

        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.freeze_panes is None
        wb.close()

    def test_estonian_characters(self, tmp_path: Path) -> None:
        """Should preserve Estonian characters (õ, ä, ö, ü)."""
        csv_file = tmp_path / "test.csv"
        estonian_text = "Muusika õpetaja käib jõulul ääremaale."
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([estonian_text])

        excel_file = csv_to_excel(csv_file)

        wb = load_workbook(excel_file)
        ws = wb.active
        assert ws.cell(1, 1).value == estonian_text
        wb.close()

    def test_header_rows_bold(self, tmp_path: Path) -> None:
        """Header rows should be bold."""
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Header1", "Header2"])
            writer.writerow(["Header3", "Header4"])
            writer.writerow(["Header5", "Header6"])
            writer.writerow(["Data1", "Data2"])

        excel_file = csv_to_excel(csv_file, freeze_header_rows=3)

        wb = load_workbook(excel_file)
        ws = wb.active

        # First 3 rows should be bold
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(2, 1).font.bold is True
        assert ws.cell(3, 1).font.bold is True

        # Data row should not be bold
        assert ws.cell(4, 1).font.bold is False

        wb.close()

    def test_auto_width_adjustment(self, tmp_path: Path) -> None:
        """Should auto-adjust column widths."""
        csv_file = tmp_path / "test.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Short", "This is a much longer column value"])

        excel_file = csv_to_excel(csv_file, auto_width=True)

        wb = load_workbook(excel_file)
        ws = wb.active

        # Column B should be wider than column A
        col_a_width = ws.column_dimensions["A"].width
        col_b_width = ws.column_dimensions["B"].width
        assert col_b_width > col_a_width

        wb.close()

    def test_empty_csv_raises_error(self, tmp_path: Path) -> None:
        """Empty CSV should raise ValueError."""
        csv_file = tmp_path / "empty.csv"
        csv_file.touch()

        with pytest.raises(ValueError, match="CSV file is empty"):
            csv_to_excel(csv_file)

    def test_missing_csv_raises_error(self, tmp_path: Path) -> None:
        """Missing CSV should raise FileNotFoundError."""
        csv_file = tmp_path / "nonexistent.csv"

        with pytest.raises(FileNotFoundError, match="CSV file not found"):
            csv_to_excel(csv_file)

    def test_muis_89_columns(self, tmp_path: Path) -> None:
        """Should handle 89-column MUIS format."""
        csv_file = tmp_path / "muis.csv"
        with open(csv_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            # 3 header rows
            writer.writerow([f"Meta{i}" for i in range(89)])
            writer.writerow([f"Col{i}" for i in range(89)])
            writer.writerow([f"Val{i}" for i in range(89)])
            # Data row
            writer.writerow([f"Data{i}" for i in range(89)])

        excel_file = csv_to_excel(csv_file, freeze_header_rows=3)

        wb = load_workbook(excel_file)
        ws = wb.active

        assert ws.max_column == 89
        assert ws.max_row == 4
        assert ws.freeze_panes == "A4"

        # Check last column (CK = column 89)
        assert ws.cell(2, 89).value == "Col88"
        assert ws.cell(4, 89).value == "Data88"

        wb.close()
